"""
AI Config Router - Admin action recording for learning data.

Provides endpoint for:
- Admin actions recording (operator edits to AI replies for training data)

Note: AI configuration is managed through /settings/update endpoint.
"""

import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from fastapi import APIRouter
from pydantic import BaseModel

# Excel 操作库
try:
    from openpyxl import Workbook, load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

router = APIRouter()

# Import settings service for database-backed prompt storage
from services.settings import get_settings_service, SettingCategory
from utils.path_utils import get_project_root

# Settings file paths
SETTINGS_DIR = get_project_root() / "settings"
ADMIN_ACTIONS_EXCEL = SETTINGS_DIR / "admin_actions.xlsx"  # 操作记录保存到 Excel


# ============================================================================
# Pydantic Models
# ============================================================================


class RecordAdminActionRequest(BaseModel):
    """Request to record an admin action."""

    message_id: str
    action_type: str  # EDIT, CANCEL, APPROVE
    original_content: str = ""
    modified_content: Optional[str] = None
    reason: Optional[str] = None
    admin_id: Optional[str] = None
    serial: Optional[str] = None
    customer_name: Optional[str] = None
    # 新增字段
    customer_id: Optional[str] = None  # 客户ID
    customer_message: Optional[str] = None  # 客户最新消息（问题）
    context: Optional[List[Dict]] = None  # 对话上下文

# ============================================================================
# File Persistence Helpers
# ============================================================================

# 注意：admin_actions.json 已废弃，所有操作记录现在只保存到 admin_actions.xlsx


def save_admin_action_to_excel(action: Dict) -> bool:
    """
    保存单条操作记录到 Excel 文件。

    只保存 EDIT 类型的操作（用户修改了 AI 回复）。
    上下文从数据库获取。

    Args:
        action: 操作记录字典

    Returns:
        是否保存成功
    """
    if not OPENPYXL_AVAILABLE:
        return False

    action_type = action.get("action_type", "").upper()
    
    # 只保存 EDIT 类型
    if action_type != "EDIT":
        return False
    
    # 必须有修改后的内容
    if not action.get("modified_content"):
        return False

    try:
        SETTINGS_DIR.mkdir(parents=True, exist_ok=True)

        # 从数据库获取上下文
        context = []
        customer_message = ""
        customer_id = action.get("customer_name", "")

        try:
            from wecom_automation.database.schema import get_connection, get_db_path

            db_path = get_db_path(None)
            if db_path.exists():
                conn = get_connection(str(db_path))
                cursor = conn.cursor()

                # 根据 customer_name 和 serial 查找客户
                customer_name = action.get("customer_name", "")
                serial = action.get("serial", "")

                if customer_name:
                    # 查找客户 ID
                    cursor.execute(
                        """
                        SELECT c.id, c.name
                        FROM customers c
                        JOIN kefus k ON c.kefu_id = k.id
                        JOIN kefu_devices kd ON k.id = kd.kefu_id
                        JOIN devices d ON kd.device_id = d.id
                        WHERE c.name = ? AND (d.serial = ? OR ? = '')
                        ORDER BY c.updated_at DESC
                        LIMIT 1
                    """,
                        (customer_name, serial, serial),
                    )

                    customer_row = cursor.fetchone()

                    if customer_row:
                        customer_db_id = customer_row["id"]

                        # 获取最近 5 条消息作为上下文
                        cursor.execute(
                            """
                            SELECT content, is_from_kefu, timestamp_raw
                            FROM messages
                            WHERE customer_id = ?
                            ORDER BY 
                                CASE WHEN ui_position IS NOT NULL THEN 0 ELSE 1 END,
                                ui_position DESC,
                                COALESCE(timestamp_parsed, created_at) DESC
                            LIMIT 5
                        """,
                            (customer_db_id,),
                        )

                        messages = cursor.fetchall()

                        # 逆序，让最旧的在前
                        messages = list(reversed(messages))

                        for msg in messages:
                            context.append(
                                {
                                    "content": msg["content"] or "",
                                    "is_from_kefu": bool(msg["is_from_kefu"]),
                                }
                            )

                        # 最后一条客户消息作为 "问题"
                        for msg in reversed(messages):
                            if not msg["is_from_kefu"]:
                                customer_message = msg["content"] or ""
                                break

                conn.close()
        except Exception as e:
            print(f"Failed to get context from database: {e}")

        # 创建或加载工作簿
        if ADMIN_ACTIONS_EXCEL.exists():
            wb = load_workbook(ADMIN_ACTIONS_EXCEL)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "操作记录"
            # 写入表头
            headers = [
                "序号",
                "分类",
                "发送者",
                "上文1",
                "发送者",
                "上文2",
                "发送者",
                "上文3",
                "发送者",
                "上文4",
                "发送者",
                "文本消息内容",
                "确定回复",
            ]
            ws.append(headers)

        # 构建行数据
        row_num = ws.max_row  # 序号
        customer_name = action.get("customer_name", "")

        row = [
            row_num,  # 序号
            "",  # 分类（留空，由人工标注）
        ]

        # 添加上下文（最多4条）
        for i in range(4):
            if i < len(context):
                msg = context[i]
                is_from_kefu = msg.get("is_from_kefu", False)
                sender = "客服" if is_from_kefu else customer_name
                content = msg.get("content", "")
            else:
                sender = ""
                content = ""
            row.extend([sender, content])

        # 问题发送者 (客户名称)
        row.append(customer_name)
        # 问题 = 客户最新消息
        row.append(customer_message)
        # 回复 = 用户修改后的内容
        row.append(action.get("modified_content", ""))

        ws.append(row)
        wb.save(ADMIN_ACTIONS_EXCEL)
        return True

    except Exception as e:
        print(f"Failed to save admin action to Excel: {e}")
        return False



# ============================================================================
# API Endpoints - Admin Actions
# ============================================================================


@router.post("/admin-action")
async def record_admin_action(request: RecordAdminActionRequest):
    """
    Record an operator action on an AI reply.

    This is called when an operator edits, cancels, or approves an AI-generated message.

    - EDIT: 保存到 Excel（包含修改内容，用于训练数据）
    - CANCEL: 不保存
    - APPROVE: 不保存
    """
    action_type = request.action_type.upper()

    # 构建操作记录
    action_dict = {
        "id": str(uuid.uuid4()),
        "message_id": request.message_id,
        "action_type": action_type,
        "original_content": request.original_content,
        "modified_content": request.modified_content,
        "reason": request.reason,
        "admin_id": request.admin_id,
        "serial": request.serial,
        "customer_name": request.customer_name,
        "customer_id": request.customer_id,
        "customer_message": request.customer_message,
        "context": request.context,
        "created_at": datetime.now().isoformat(),
    }

    # 只有 EDIT 类型保存到 Excel
    excel_saved = False
    if action_type == "EDIT":
        excel_saved = save_admin_action_to_excel(action_dict)

    return {"success": True, "action_id": action_dict["id"], "excel_saved": excel_saved}
